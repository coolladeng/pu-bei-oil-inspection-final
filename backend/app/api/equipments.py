"""设备管理API"""
import json
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.equipment import Equipment, EquipCheckItem
from app.models.dept import SysDept
from app.schemas.equipment import (
    EquipmentCreate, EquipmentUpdate, EquipmentResponse,
    CheckItemCreate, CheckItemUpdate, CheckItemResponse,
)

router = APIRouter()


@router.get("")
async def list_equipments(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    name: str | None = None,
    code: str | None = None,
    dept_id: int | None = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(Equipment).options(selectinload(Equipment.dept))
    if name:
        query = query.where(Equipment.name.like(f"%{name}%"))
    if code:
        query = query.where(Equipment.code.like(f"%{code}%"))
    if dept_id:
        query = query.where(Equipment.dept_id == dept_id)

    total_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(total_query)).scalar() or 0

    query = query.order_by(Equipment.id.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    equipments = result.scalars().all()

    return {
        "list": [{
            "id": e.id, "code": e.code, "name": e.name,
            "deptId": e.dept_id,
            "deptName": e.dept.name if e.dept else None,
            "modelNo": e.model_no, "manufacturer": e.manufacturer,
            "location": e.location, "category": e.category,
            "status": e.status,
        } for e in equipments],
        "total": total, "page": page, "pageSize": page_size,
    }


@router.post("")
async def create_equipment(data: EquipmentCreate, db: AsyncSession = Depends(get_db)):
    existing = await db.execute(select(Equipment).where(Equipment.code == data.code))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="设备编号已存在")
    equip = Equipment(**data.model_dump())
    db.add(equip)
    await db.commit()
    await db.refresh(equip)
    return {"id": equip.id, "message": "创建成功"}


@router.put("/{equip_id}")
async def update_equipment(equip_id: int, data: EquipmentUpdate, db: AsyncSession = Depends(get_db)):
    equip = await db.get(Equipment, equip_id)
    if not equip:
        raise HTTPException(status_code=404, detail="设备不存在")
    update_data = data.model_dump(exclude_none=True)
    if "code" in update_data and update_data["code"] != equip.code:
        existing = await db.execute(select(Equipment).where(Equipment.code == update_data["code"]))
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="设备编号已存在")
    for key, val in update_data.items():
        setattr(equip, key, val)
    await db.commit()
    return {"message": "更新成功"}


@router.delete("/{equip_id}")
async def delete_equipment(equip_id: int, db: AsyncSession = Depends(get_db)):
    equip = await db.get(Equipment, equip_id)
    if not equip:
        raise HTTPException(status_code=404, detail="设备不存在")
    await db.delete(equip)
    await db.commit()
    return {"message": "删除成功"}


@router.get("/{equip_id}/qrcode")
async def generate_qrcode(equip_id: int, db: AsyncSession = Depends(get_db)):
    equip = await db.get(Equipment, equip_id)
    if not equip:
        raise HTTPException(status_code=404, detail="设备不存在")
    return {"qrcode": f"equip://{equip.id}?code={equip.code}"}


# ---- 检查项 CRUD ----

@router.get("/{equip_id}/check-items")
async def list_check_items(equip_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(EquipCheckItem).where(EquipCheckItem.equip_id == equip_id).order_by(EquipCheckItem.sort_order)
    )
    items = result.scalars().all()
    return [{
        "id": i.id, "equipId": i.equip_id,
        "itemName": i.item_name, "itemType": i.item_type,
        "options": i.options, "normalMin": i.normal_min,
        "normalMax": i.normal_max, "unit": i.unit,
        "sortOrder": i.sort_order, "required": i.required,
    } for i in items]


@router.post("/{equip_id}/check-items")
async def create_check_item(equip_id: int, data: CheckItemCreate, db: AsyncSession = Depends(get_db)):
    equip = await db.get(Equipment, equip_id)
    if not equip:
        raise HTTPException(status_code=404, detail="设备不存在")
    item = EquipCheckItem(equip_id=equip_id, **data.model_dump())
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "message": "创建成功"}


@router.put("/{equip_id}/check-items/{item_id}")
async def update_check_item(equip_id: int, item_id: int, data: CheckItemUpdate, db: AsyncSession = Depends(get_db)):
    item = await db.get(EquipCheckItem, item_id)
    if not item or item.equip_id != equip_id:
        raise HTTPException(status_code=404, detail="检查项不存在")
    update_data = data.model_dump(exclude_none=True)
    for key, val in update_data.items():
        setattr(item, key, val)
    await db.commit()
    return {"message": "更新成功"}


@router.delete("/{equip_id}/check-items/{item_id}")
async def delete_check_item(equip_id: int, item_id: int, db: AsyncSession = Depends(get_db)):
    item = await db.get(EquipCheckItem, item_id)
    if not item or item.equip_id != equip_id:
        raise HTTPException(status_code=404, detail="检查项不存在")
    await db.delete(item)
    await db.commit()
    return {"message": "删除成功"}


# ---- Excel 导入 ----

@router.post("/import")
async def import_equipments(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(await file.read())
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="工作簿为空")
        created, skipped = 0, 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or (len(row) == 0) or (row[0] is None and row[1] is None):
                continue
            vals = list(row) + [None] * 7
            code, name, model_no, manufacturer, dept_name, location, category = vals[:7]

            if not code or not name:
                errors.append(f"第{row_idx}行: 编号和名称为必填")
                continue

            # 检查重复
            existing = await db.execute(select(Equipment).where(Equipment.code == str(code)))
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            # 尝试通过部门名称查找 dept_id
            dept_id = None
            if dept_name:
                dept_result = await db.execute(
                    select(SysDept).where(SysDept.name == str(dept_name))
                )
                dept = dept_result.scalar_one_or_none()
                if dept:
                    dept_id = dept.id

            equip = Equipment(
                code=str(code),
                name=str(name),
                model_no=str(model_no) if model_no else None,
                manufacturer=str(manufacturer) if manufacturer else None,
                location=str(location) if location else None,
                category=str(category) if category else None,
                dept_id=dept_id or 0,
                status=1,
            )
            db.add(equip)
            created += 1

        await db.commit()
        msg = f"导入完成：成功 {created} 条，跳过 {skipped} 条"
        if errors:
            msg += f"，{len(errors)} 条格式错误"
        return {"message": msg, "created": created, "skipped": skipped, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")


# ---- 下载导入模板 ----

@router.get("/template")
async def download_template():
    """返回 Excel 导入模板"""
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "设备导入模板"
    headers = ["编号", "名称", "型号", "生产厂家", "所属部门", "安装位置", "设备类别"]
    ws.append(headers)
    # 示例数据
    ws.append(["EQ-001", "离心泵A", "IS80-65-160", "大庆油田装备", "大队A", "1号联合站", "泵类"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=equipment_template.xlsx"},
    )
